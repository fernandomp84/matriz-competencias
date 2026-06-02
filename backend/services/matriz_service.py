"""
Servicio generador de Matriz_Competencias.csv a partir de JSON.
Adaptado de generar_matriz_competencias.py para funcionar con FastAPI.
"""

import base64
import csv
import io
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Utilidades ─────────────────────────────────────────────────────────────────

def limpiar(v) -> str:
    if not v:
        return ""
    t = str(v).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r" {2,}", " ", t).strip()


def truncar(v, max_len: int = 80) -> str:
    txt = limpiar(v)
    return txt[:max_len] if len(txt) > max_len else txt


def html_desc(texto: str) -> str:
    return f'<p dir="ltr" style="text-align:left;">{texto}</p>'


def extraer_padre_meta(meta_cod: str) -> str:
    partes = meta_cod.strip().split("_")
    if len(partes) > 1 and re.match(r"^M\d+$", partes[0]):
        return "_".join(partes[1:])
    return partes[-1]


# ── Generación de filas ─────────────────────────────────────────────────────────

def generar_filas(data: dict, tm: dict | None = None) -> tuple[list[dict], list[dict]]:
    """
    tm: mapa {codigo_padre: descripcion} leído de la hoja TM.
    Si existe entrada en TM para el prog_id, usa esa descripción como
    Nombre_corto y Descripción del padre principal.
    Si no, conserva el comportamiento anterior (programa.nombre o prog_id).
    """
    tm = tm or {}
    filas: list[dict] = []
    seen: set[str] = set()
    errores: list[dict] = []

    def agregar(id_paterno: str, id_nodo: str, nombre_corto: str, desc: str):
        if id_nodo in seen:
            return
        seen.add(id_nodo)
        filas.append({
            "Número ID paterno": id_paterno,
            "Número ID": id_nodo,
            "Nombre_corto": nombre_corto,
            "Descripción": html_desc(desc),
            "Formato de descripción": 1,
        })

    def err(id_siga: str, campo: str, mensaje: str):
        errores.append({"idCursoSiga": id_siga, "campo": campo, "error": mensaje})

    for id_siga, curso in data.items():
        for idx, comp in enumerate(curso.get("competencias", []), 1):
            try:
                prog = comp.get("programa", {})
                meta = comp.get("metaCompetencia", {})
                ra = comp.get("resultadoAprendizaje", {})
                ind = comp.get("indicadorDesempeno", {})

                meta_cod = limpiar(meta.get("codigo", ""))
                if not meta_cod:
                    err(id_siga, f"competencias[{idx}].metaCompetencia.codigo",
                        "Código de metacompetencia vacío — fila ignorada")
                    continue

                ra_cod = limpiar(ra.get("codigo", ""))
                if not ra_cod:
                    err(id_siga, f"competencias[{idx}].resultadoAprendizaje.codigo",
                        "Código de resultado de aprendizaje vacío — fila ignorada")
                    continue

                ind_cod = limpiar(ind.get("codigo", ""))
                if not ind_cod:
                    err(id_siga, f"competencias[{idx}].indicadorDesempeno.codigo",
                        "Código de indicador de desempeño vacío — fila ignorada")
                    continue

                # Meta SABE: el padre principal siempre es "SABER_UX" (regla fija).
                # Para el resto se deriva del propio meta_cod.
                if meta_cod == "SABE":
                    prog_id = "SABER_UX"
                else:
                    prog_id = extraer_padre_meta(meta_cod)

                # Prioridad: hoja TM → programa.nombre del JSON → prog_id como fallback
                prog_nombre = (
                    tm.get(prog_id)
                    or limpiar(prog.get("nombre", ""))
                    or prog_id
                )
                agregar("", prog_id, truncar(prog_nombre, 80), prog_nombre)

                meta_titulo = limpiar(meta.get("titulo", "")) or meta_cod
                agregar(prog_id, meta_cod, truncar(meta_titulo, 80), meta_titulo)

                ra_desc = limpiar(ra.get("descripcion", "")) or ra_cod
                agregar(meta_cod, ra_cod, truncar(ra_desc, 80), ra_desc)

                ind_desc = limpiar(ind.get("descripcion", "")) or ind_cod
                agregar(ra_cod, ind_cod, truncar(ind_desc, 80), ind_desc)

            except Exception as exc:
                err(id_siga, f"competencias[{idx}]", f"Error inesperado: {exc}")

    return filas, errores


# ── CSV ─────────────────────────────────────────────────────────────────────────

CAMPOS_CSV = [
    "Número ID paterno",
    "Número ID",
    "Nombre_corto",
    "Descripción",
    "Formato de descripción",
]


def generar_csv_string(filas: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(
        output, fieldnames=CAMPOS_CSV, delimiter=";", quoting=csv.QUOTE_ALL
    )
    writer.writeheader()
    writer.writerows(filas)
    return output.getvalue()


# ── XLSX de errores ─────────────────────────────────────────────────────────────

def generar_errores_xlsx_b64(errores: list[dict]) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errores"

    thin = Side(style="thin", color="BFBFBF")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="C00000")
    hdr_f = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["ID Curso SIGA", "Campo", "Error"]
    widths = [18, 40, 80]
    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_f
        c.fill = hdr_fill
        c.alignment = hdr_al
        c.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", start_color="FFF2F2")

    for ri, e in enumerate(errores, 2):
        ws.row_dimensions[ri].height = 35
        fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for ci, key in enumerate(["idCursoSiga", "campo", "error"], 1):
            c = ws.cell(row=ri, column=ci, value=e.get(key, ""))
            c.font = data_font
            c.fill = fill
            c.alignment = data_align
            c.border = brd

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(errores) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


# ── Servicio principal ──────────────────────────────────────────────────────────

def procesar(json_data: dict) -> dict:
    data = json_data.get("data", json_data)
    tm   = json_data.get("tm", {})
    filas, errores = generar_filas(data, tm)
    csv_content = generar_csv_string(filas)
    errores_xlsx_b64 = generar_errores_xlsx_b64(errores)
    total_prog = sum(1 for f in filas if not f["Número ID paterno"])
    return {
        "filas": filas,
        "errores": errores,
        "csvContent": csv_content,
        "erroresXlsxB64": errores_xlsx_b64,
        "stats": {
            "totalFilas": len(filas),
            "programas": total_prog,
            "totalErrores": len(errores),
            "tmEntradas": len(tm),
        },
    }
