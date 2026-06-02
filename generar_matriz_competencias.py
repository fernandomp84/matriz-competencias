"""
Generador de Matriz_Competencias.csv a partir de salida.json

Jerarquía:
  Programa → MetaCompetencia → ResultadoAprendizaje → IndicadorDesempeno

Salida : Matriz_Competencias.csv  (UTF-8 con BOM, separador ;)
Errores: errorM.xlsx              (en la carpeta de salida)
"""

import csv
import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from dotenv import load_dotenv


# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

def cargar_configuracion() -> dict:
    load_dotenv()
    claves = {
        "input_json_path":    "OUTPUT_JSON_PATH",      # reutiliza variable existente
        "output_matrix_path": "OUTPUT_MATRIX_PATH",    # nueva variable
    }
    config = {}
    faltantes = []
    for clave, var in claves.items():
        valor = os.getenv(var, "").strip()
        if not valor:
            faltantes.append(var)
        config[clave] = valor
    if faltantes:
        raise ValueError(
            f"Variables de entorno requeridas no encontradas: {', '.join(faltantes)}"
        )
    return config


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def limpiar(v) -> str:
    if not v:
        return ""
    t = str(v).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r" {2,}", " ", t).strip()


def truncar(v, max_len: int = 80) -> str:
    txt = limpiar(v)
    return txt[:max_len] if len(txt) > max_len else txt


def html_desc(texto: str) -> str:
    return f'<p dir="ltr" style="text-align:left;">{texto}</p>'


def extraer_padre_meta(meta_cod: str) -> str:
    """
    Extrae el identificador padre de una metacompetencia.
    Si el primer segmento coincide con M+dígitos, lo omite y retorna el resto.
      Ej: 'M012_UX_INS'  → 'UX_INS'
          'M013_ELECTIVA' → 'ELECTIVA'
    Fallback (sin patrón M+dígitos): retorna la última parte.
    """
    partes = meta_cod.strip().split("_")
    if len(partes) > 1 and re.match(r"^M\d+$", partes[0]):
        return "_".join(partes[1:])
    return partes[-1]


# ---------------------------------------------------------------------------
# Leer JSON
# ---------------------------------------------------------------------------

def leer_json(ruta_json: str) -> dict:
    with open(ruta_json, "rb") as f:
        raw = f.read().decode("utf-8", errors="replace")
    raw = re.sub(r"^/\*.*?\*/\s*", "", raw, flags=re.DOTALL)
    raw = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", raw)
    return json.loads(raw)["data"]


# ---------------------------------------------------------------------------
# Generación de filas
# ---------------------------------------------------------------------------

def generar_filas(data: dict) -> tuple[list[dict], list[dict]]:
    """
    Recorre todos los cursos y competencias del JSON.
    Retorna (filas, errores).
    Las filas no contienen duplicados (controlado por 'seen').
    """
    filas: list[dict] = []
    seen: set[str] = set()         # IDs ya registrados
    errores: list[dict] = []

    def agregar(id_paterno: str, id_nodo: str, nombre_corto: str, desc: str):
        if id_nodo in seen:
            return
        seen.add(id_nodo)
        filas.append({
            "Número ID paterno":       id_paterno,
            "Número ID":               id_nodo,
            "Nombre_corto":            nombre_corto,
            "Descripción":             html_desc(desc),
            "Formato de descripción":  1,
        })

    def err(id_siga: str, campo: str, mensaje: str):
        errores.append({
            "idCursoSiga": id_siga,
            "campo":       campo,
            "error":       mensaje,
        })

    for id_siga, curso in data.items():
        for idx, comp in enumerate(curso.get("competencias", []), 1):
            try:
                prog = comp.get("programa", {})
                meta = comp.get("metaCompetencia", {})
                ra   = comp.get("resultadoAprendizaje", {})
                ind  = comp.get("indicadorDesempeno", {})

                # ── Validaciones críticas ──────────────────────────────────
                meta_cod = limpiar(meta.get("codigo", ""))
                if not meta_cod:
                    err(id_siga, f"competencias[{idx}].metaCompetencia.codigo",
                        "Código de metacompetencia vacío — fila ignorada")
                    continue

                ra_cod = limpiar(ra.get("codigo", ""))
                if not ra_cod:
                    err(id_siga, f"competencias[{idx}].resultadoAprendizaje.codigo",
                        "Código de resultado de aprendizaje vacío — fila ignorada")
                    continue

                ind_cod = limpiar(ind.get("codigo", ""))
                if not ind_cod:
                    err(id_siga, f"competencias[{idx}].indicadorDesempeno.codigo",
                        "Código de indicador de desempeño vacío — fila ignorada")
                    continue

                # ── 1. PROGRAMA (raíz, sin padre) ──────────────────────────
                prog_id     = extraer_padre_meta(meta_cod)    # e.g. UX_INS, ELECTIVA
                prog_nombre = limpiar(prog.get("nombre", "")) or prog_id
                agregar(
                    id_paterno   = "",
                    id_nodo      = prog_id,
                    nombre_corto = truncar(prog_nombre, 80),
                    desc         = prog_nombre,
                )

                # ── 2. METACOMPETENCIA (padre = programa) ──────────────────
                meta_titulo = limpiar(meta.get("titulo", "")) or meta_cod
                agregar(
                    id_paterno   = prog_id,
                    id_nodo      = meta_cod,
                    nombre_corto = truncar(meta_titulo, 80),
                    desc         = meta_titulo,
                )

                # ── 3. RESULTADO DE APRENDIZAJE (padre = metacompetencia) ──
                ra_desc = limpiar(ra.get("descripcion", "")) or ra_cod
                agregar(
                    id_paterno   = meta_cod,
                    id_nodo      = ra_cod,
                    nombre_corto = truncar(ra_desc, 80),
                    desc         = ra_desc,
                )

                # ── 4. INDICADOR DE DESEMPEÑO (padre = resultado de aprendizaje)
                ind_desc = limpiar(ind.get("descripcion", "")) or ind_cod
                agregar(
                    id_paterno   = ra_cod,
                    id_nodo      = ind_cod,
                    nombre_corto = truncar(ind_desc, 80),
                    desc         = ind_desc,
                )

            except Exception as exc:
                err(id_siga, f"competencias[{idx}]",
                    f"Error inesperado: {exc}")

    return filas, errores


# ---------------------------------------------------------------------------
# Escritura CSV
# ---------------------------------------------------------------------------

CAMPOS_CSV = [
    "Número ID paterno",
    "Número ID",
    "Nombre_corto",
    "Descripción",
    "Formato de descripción",
]


def escribir_csv(filas: list[dict], ruta_csv: str) -> None:
    with open(ruta_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=CAMPOS_CSV,
            delimiter=";",
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        writer.writerows(filas)


# ---------------------------------------------------------------------------
# Escritura errorM.xlsx
# ---------------------------------------------------------------------------

def escribir_errores_xlsx(errores: list[dict], ruta_xlsx: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errores"

    thin   = Side(style="thin", color="BFBFBF")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_f  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", start_color="C00000")
    hdr_al = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["ID Curso SIGA", "Campo", "Error"]
    widths  = [18, 40, 80]

    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_f; c.fill = hdr_fill
        c.alignment = hdr_al; c.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    alt_fill   = PatternFill("solid", start_color="FFF2F2")

    for ri, e in enumerate(errores, 2):
        ws.row_dimensions[ri].height = 35
        fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for ci, key in enumerate(["idCursoSiga", "campo", "error"], 1):
            c = ws.cell(row=ri, column=ci, value=e.get(key, ""))
            c.font = data_font; c.fill = fill
            c.alignment = data_align; c.border = brd

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(errores) + 1}"
    wb.save(ruta_xlsx)


# ---------------------------------------------------------------------------
# Orquestador
# ---------------------------------------------------------------------------

def procesar(config: dict) -> None:
    ruta_json  = Path(config["input_json_path"])  / "salida.json"
    ruta_csv   = Path(config["output_matrix_path"]) / "Matriz_Competencias.csv"
    ruta_errxl = Path(config["output_matrix_path"]) / "errorM.xlsx"

    Path(config["output_matrix_path"]).mkdir(parents=True, exist_ok=True)

    print(f"[INFO] Leyendo JSON: {ruta_json}")
    data = leer_json(str(ruta_json))
    print(f"[INFO] Cursos encontrados: {len(data)}")

    filas, errores = generar_filas(data)

    print(f"[INFO] Filas generadas (sin duplicados): {len(filas)}")
    escribir_csv(filas, str(ruta_csv))
    print(f"[OK]  CSV guardado: {ruta_csv}")

    escribir_errores_xlsx(errores, str(ruta_errxl))
    print(f"[OK]  Errores guardados: {ruta_errxl}")

    # Resumen
    total_prog  = sum(1 for f in filas if not f["Número ID paterno"])
    total_meta  = sum(1 for f in filas if f["Número ID"].startswith("M0") or "_" in f["Número ID"] and not f["Número ID"].startswith(("IND","RA")))
    print(f"\n--- Resumen ---")
    print(f"  Total filas CSV      : {len(filas)}")
    print(f"  Programas (raíz)     : {total_prog}")
    print(f"  Errores registrados  : {len(errores)}")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    try:
        cfg = cargar_configuracion()
        procesar(cfg)
    except Exception as e:
        print(f"[ERROR FATAL] {e}")
        raise
